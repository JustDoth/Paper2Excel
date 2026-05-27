"""Non-GUI core logic for converting paper Excel rows into structured fields.

This module is intentionally UI-agnostic. A desktop GUI can import these
functions to read workbooks, build model prompts, call an OpenAI-compatible
chat-completions API, parse JSON output, and write enriched Excel files.
"""

from __future__ import annotations

import json
import re
import time
import csv
from urllib.parse import quote
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence


JsonDict = dict[str, Any]
ChatCallable = Callable[[list[JsonDict]], str]


class Paper2ExcelError(Exception):
    """Base exception for this module."""


class JsonParseError(Paper2ExcelError):
    """Raised when a model response cannot be parsed into a JSON object."""


class ChatCompletionError(Paper2ExcelError):
    """Raised when a chat-completions API call fails."""


@dataclass(frozen=True)
class ColumnMapping:
    """Maps one Excel source column into a prompt input field."""

    source: str
    alias: str | None = None
    description: str = ""

    @property
    def prompt_name(self) -> str:
        return self.alias or self.source


@dataclass(frozen=True)
class OutputField:
    """Defines one structured field expected from the model."""

    name: str
    description: str
    field_type: str = "string"
    required: bool = True
    examples: Sequence[Any] = field(default_factory=tuple)


@dataclass(frozen=True)
class ChatCompletionSettings:
    """Settings for OpenAI-compatible chat-completions endpoints."""

    api_key: str
    model: str
    base_url: str = "https://api.openai.com/v1"
    temperature: float = 0.2
    timeout_seconds: int = 180
    max_tokens: int | None = 3000
    proxy_url: str = ""
    extra_headers: Mapping[str, str] = field(default_factory=dict)


def read_excel_rows(
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> list[JsonDict]:
    """Read an Excel or CSV sheet into a list of row dictionaries.

    Empty rows are skipped. Cell values are returned as openpyxl-native Python
    values such as strings, numbers, datetimes, and booleans.
    """

    source_path = Path(path)
    if source_path.suffix.lower() == ".csv":
        return read_csv_rows(source_path)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise Paper2ExcelError("Reading Excel requires the 'openpyxl' package.") from exc

    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = [cell.value for cell in sheet[header_row]]
    normalized_headers = [_normalize_header(value, index) for index, value in enumerate(headers)]

    rows: list[JsonDict] = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        if all(value is None for value in row):
            continue
        rows.append(
            {
                normalized_headers[index]: value
                for index, value in enumerate(row[: len(normalized_headers)])
            }
        )
    return rows


def read_csv_rows(path: str | Path) -> list[JsonDict]:
    """Read a CSV file into row dictionaries."""

    csv_path = Path(path)
    encodings = ("utf-8-sig", "gb18030")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with csv_path.open("r", encoding=encoding, newline="") as handle:
                reader = csv.reader(handle)
                try:
                    raw_headers = next(reader)
                except StopIteration:
                    return []
                headers = [_normalize_header(value, index) for index, value in enumerate(raw_headers)]
                rows: list[JsonDict] = []
                for raw_row in reader:
                    if all(value is None or str(value).strip() == "" for value in raw_row):
                        continue
                    padded = [*raw_row, *[""] * max(0, len(headers) - len(raw_row))]
                    rows.append({headers[index]: padded[index] for index in range(len(headers))})
                return rows
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    raise Paper2ExcelError(f"Reading CSV failed: {last_error}")


def write_excel_rows(
    path: str | Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str] | None = None,
    sheet_name: str = "Result",
) -> None:
    """Write dictionaries to an Excel workbook or CSV file."""

    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        _write_csv_rows(output_path, rows, columns)
        return

    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise Paper2ExcelError("Writing Excel requires the 'openpyxl' package.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    resolved_columns = list(columns or _collect_columns(rows))
    sheet.append(resolved_columns)
    for row_index, row in enumerate(rows, start=2):
        sheet.append([_excel_safe_value(row.get(column)) for column in resolved_columns])
        for column_index, column in enumerate(resolved_columns, start=1):
            if column.strip().lower() == "file attachments":
                _apply_file_attachment_hyperlink(sheet.cell(row=row_index, column=column_index), row.get(column))

    workbook.save(output_path)


def _write_csv_rows(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
    columns: Sequence[str] | None = None,
) -> None:
    resolved_columns = list(columns or _collect_columns(rows))
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(resolved_columns)
        for row in rows:
            writer.writerow([_excel_safe_value(row.get(column)) for column in resolved_columns])


def _apply_file_attachment_hyperlink(cell: Any, value: Any) -> None:
    target = _first_existing_file_path(value)
    if not target:
        return
    cell.hyperlink = _file_uri(target)
    cell.style = "Hyperlink"


def _first_existing_file_path(value: Any) -> Path | None:
    if value is None:
        return None
    for raw in str(value).split(";"):
        candidate = raw.strip().strip('"')
        if not candidate:
            continue
        if candidate.lower().startswith("file://"):
            candidate = candidate[7:].lstrip("/")
        path = Path(candidate)
        if path.exists() and path.is_file():
            return path.resolve()
    return None


def _file_uri(path: Path) -> str:
    normalized = path.resolve().as_posix()
    return "file:///" + quote(normalized, safe="/:")


def build_json_prompt(
    row: Mapping[str, Any],
    column_mappings: Sequence[ColumnMapping | Mapping[str, Any]],
    output_fields: Sequence[OutputField | Mapping[str, Any]],
    task_description: str = "Extract structured information from one paper record.",
) -> list[JsonDict]:
    """Build chat messages that ask the model to return one strict JSON object."""

    mappings = [_as_column_mapping(item) for item in column_mappings]
    fields = [_as_output_field(item) for item in output_fields]
    source_payload = {
        mapping.prompt_name: {
            "value": row.get(mapping.source),
            "source_column": mapping.source,
            "description": mapping.description,
        }
        for mapping in mappings
    }
    output_schema = {
        field.name: {
            "type": field.field_type,
            "required": field.required,
            "description": field.description,
            "examples": list(field.examples),
        }
        for field in fields
    }
    required_names = [field.name for field in fields if field.required]

    system = (
        "You convert tabular records into structured JSON. "
        "Use only the provided row content. Do not invent unsupported facts. "
        "Return only one valid JSON object, with no markdown, comments, or extra text."
    )
    user = {
        "task": task_description,
        "input_columns": source_payload,
        "output_fields": output_schema,
        "rules": [
            "Use null when the answer is unavailable or cannot be inferred.",
            "Do not invent facts that are not supported by the input.",
            "Return all required keys exactly as named.",
            "Use strings for long text fields; use numbers only when the field asks for a number.",
            "If a quote is requested, quote only text that appears in the input.",
        ],
        "required_keys": required_names,
    }
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": json.dumps(user, ensure_ascii=False, default=str)},
    ]


def call_chat_completion(
    messages: list[JsonDict],
    settings: ChatCompletionSettings,
) -> str:
    """Call an OpenAI-compatible /chat/completions endpoint and return content."""

    url = build_chat_completions_url(settings.base_url)
    payload: JsonDict = {
        "model": settings.model,
        "messages": messages,
        "temperature": settings.temperature,
    }
    if settings.max_tokens is not None:
        payload["max_tokens"] = settings.max_tokens

    headers = {"Content-Type": "application/json", **dict(settings.extra_headers)}
    if settings.api_key.strip():
        headers["Authorization"] = f"Bearer {settings.api_key.strip()}"

    response_payload = _post_json(url, headers, payload, settings)

    try:
        return response_payload["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as exc:
        raise ChatCompletionError(f"Unexpected chat completion response: {response_payload}") from exc


def _post_json(
    url: str,
    headers: Mapping[str, str],
    payload: Mapping[str, Any],
    settings: ChatCompletionSettings,
) -> JsonDict:
    """POST JSON with requests when available, falling back to urllib."""

    try:
        import requests
    except ImportError as import_error:
        return _post_json_with_urllib(url, headers, payload, settings, import_error)

    try:
        proxies = None
        if settings.proxy_url.strip():
            proxy = normalize_proxy_url(settings.proxy_url)
            proxies = {"http": proxy, "https": proxy}
        response = requests.post(
            url,
            headers=dict(headers),
            json=payload,
            timeout=settings.timeout_seconds,
            proxies=proxies,
        )
        if response.status_code >= 400:
            raise ChatCompletionError(
                f"Chat completion failed with HTTP {response.status_code}: {response.text}"
            )
        return response.json()
    except ChatCompletionError:
        raise
    except (requests.RequestException, TimeoutError, json.JSONDecodeError) as exc:
        raise ChatCompletionError(f"Chat completion request failed: {exc}") from exc


def _post_json_with_urllib(
    url: str,
    headers: Mapping[str, str],
    payload: Mapping[str, Any],
    settings: ChatCompletionSettings,
    import_error: ImportError,
) -> JsonDict:
    """Fallback HTTP client using the standard library."""

    import urllib.error
    import urllib.request

    request = urllib.request.Request(
        url=url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=dict(headers),
        method="POST",
    )
    try:
        if settings.proxy_url.strip():
            proxy = normalize_proxy_url(settings.proxy_url)
            opener = urllib.request.build_opener(urllib.request.ProxyHandler({"http": proxy, "https": proxy}))
            response_context = opener.open(request, timeout=settings.timeout_seconds)
        else:
            response_context = urllib.request.urlopen(request, timeout=settings.timeout_seconds)
        with response_context as response:
            return json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise ChatCompletionError(f"Chat completion failed with HTTP {exc.code}: {detail}") from exc
    except Exception as exc:
        raise ChatCompletionError(
            "Chat completion request failed. "
            f"Requests import failed first ({type(import_error).__name__}: {import_error}); "
            f"urllib fallback failed with {type(exc).__name__}: {exc}"
        ) from exc


def build_chat_completions_url(base_url: str) -> str:
    """Build the final chat-completions URL from a base URL or full endpoint."""

    value = base_url.strip().rstrip("/")
    if not value:
        raise ChatCompletionError("Base URL is empty.")
    if not re.match(r"^https?://", value, flags=re.IGNORECASE):
        raise ChatCompletionError("Base URL must start with http:// or https://.")
    if value.endswith("/chat/completions"):
        return value
    return f"{value}/chat/completions"


def normalize_proxy_url(proxy_url: str) -> str:
    """Accept localhost:7890 or http://localhost:7890 style proxy values."""

    value = proxy_url.strip()
    if not value:
        return ""
    if re.match(r"^https?://", value, flags=re.IGNORECASE):
        return value
    return f"http://{value}"


def optimize_field_prompt(
    column_name: str,
    user_description: str,
    settings: ChatCompletionSettings,
    output_language: str = "English",
) -> str:
    """Rewrite a plain-language column request into a stable field prompt."""

    messages = [
        {
            "role": "system",
            "content": (
                "You are a prompt engineer for batch Excel document analysis. "
                "Return only the improved field-level instruction, no markdown."
            ),
        },
        {
            "role": "user",
            "content": (
                f"Excel output column name: {column_name}\n"
                f"User plain-language request: {user_description}\n\n"
                "Rewrite it as one stable instruction for an AI that analyzes one Excel row at a time.\n"
                f"Output language should be {output_language} unless the user clearly asks otherwise.\n"
                "The instruction must define length/format when possible, forbid unsupported claims, "
                "and state what to return when evidence is insufficient."
            ),
        },
    ]
    response = call_chat_completion(messages, settings)
    return response.strip().strip("`").strip()


def parse_json_object(text: str, required_keys: Iterable[str] = ()) -> JsonDict:
    """Parse a JSON object from raw model output.

    The parser accepts a bare JSON object or a fenced markdown JSON block.
    Required keys are validated after parsing.
    """

    candidates = [text]
    fenced_match = re.search(r"```(?:json)?\s*(.*?)```", text, flags=re.IGNORECASE | re.DOTALL)
    if fenced_match:
        candidates.insert(0, fenced_match.group(1))

    object_match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if object_match:
        candidates.append(object_match.group(0))

    last_error: Exception | None = None
    for candidate in candidates:
        try:
            value = json.loads(candidate.strip())
        except json.JSONDecodeError as exc:
            last_error = exc
            continue
        if not isinstance(value, dict):
            raise JsonParseError("Model response JSON must be an object.")
        missing = [key for key in required_keys if key not in value]
        if missing:
            raise JsonParseError(f"Model response missing required keys: {', '.join(missing)}")
        return value

    raise JsonParseError(f"Could not parse JSON object from model response: {last_error}")


def process_excel(
    input_path: str | Path,
    output_path: str | Path,
    column_mappings: Sequence[ColumnMapping | Mapping[str, Any]],
    output_fields: Sequence[OutputField | Mapping[str, Any]],
    chat: ChatCallable | None = None,
    settings: ChatCompletionSettings | None = None,
    sheet_name: str | None = None,
    header_row: int = 1,
    task_description: str = "Extract structured information from one paper record.",
    max_parse_retries: int = 2,
    retry_sleep_seconds: float = 0.5,
) -> list[JsonDict]:
    """Read an Excel file, enrich each row with model JSON, and write a new workbook.

    Pass either a custom ``chat(messages) -> str`` callable for tests/offline use
    or ``settings`` for a real OpenAI-compatible API call.
    """

    if chat is None:
        if settings is None:
            raise ValueError("Either 'chat' or 'settings' must be provided.")
        chat = lambda messages: call_chat_completion(messages, settings)

    fields = [_as_output_field(item) for item in output_fields]
    required_keys = [field.name for field in fields if field.required]
    input_rows = read_excel_rows(input_path, sheet_name=sheet_name, header_row=header_row)
    output_rows: list[JsonDict] = []

    for row_index, row in enumerate(input_rows, start=1):
        messages = build_json_prompt(row, column_mappings, fields, task_description)
        parsed = _call_and_parse_with_retries(
            chat=chat,
            messages=messages,
            required_keys=required_keys,
            max_parse_retries=max_parse_retries,
            retry_sleep_seconds=retry_sleep_seconds,
        )
        output_rows.append({**row, **{field.name: parsed.get(field.name) for field in fields}})

    source_columns = list(input_rows[0].keys()) if input_rows else []
    output_columns = source_columns + [field.name for field in fields if field.name not in source_columns]
    write_excel_rows(output_path, output_rows, columns=output_columns)
    return output_rows


def _call_and_parse_with_retries(
    chat: ChatCallable,
    messages: list[JsonDict],
    required_keys: Sequence[str],
    max_parse_retries: int,
    retry_sleep_seconds: float,
) -> JsonDict:
    current_messages = list(messages)
    last_error: JsonParseError | None = None
    for attempt in range(max_parse_retries + 1):
        response = chat(current_messages)
        try:
            return parse_json_object(response, required_keys=required_keys)
        except JsonParseError as exc:
            last_error = exc
            if attempt >= max_parse_retries:
                break
            current_messages = [
                *messages,
                {"role": "assistant", "content": response},
                {
                    "role": "user",
                    "content": (
                        "The previous response was not valid for this task: "
                        f"{exc}. Return only one corrected JSON object."
                    ),
                },
            ]
            time.sleep(retry_sleep_seconds)
    raise last_error or JsonParseError("Failed to parse model JSON response.")


def _as_column_mapping(value: ColumnMapping | Mapping[str, Any]) -> ColumnMapping:
    if isinstance(value, ColumnMapping):
        return value
    return ColumnMapping(
        source=str(value["source"]),
        alias=value.get("alias"),
        description=str(value.get("description", "")),
    )


def _as_output_field(value: OutputField | Mapping[str, Any]) -> OutputField:
    if isinstance(value, OutputField):
        return value
    return OutputField(
        name=str(value["name"]),
        description=str(value.get("description", "")),
        field_type=str(value.get("field_type", value.get("type", "string"))),
        required=bool(value.get("required", True)),
        examples=tuple(value.get("examples", ())),
    )


def _normalize_header(value: Any, index: int) -> str:
    if value is None or str(value).strip() == "":
        return f"column_{index + 1}"
    return str(value).strip()


def _collect_columns(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def _excel_safe_value(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value
