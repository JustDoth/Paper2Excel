"""Batch execution utilities used by the desktop app."""

from __future__ import annotations

import json
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence

from .core import (
    ChatCallable,
    ChatCompletionError,
    ChatCompletionSettings,
    ColumnMapping,
    JsonParseError,
    OutputField,
    build_json_prompt,
    call_chat_completion,
    parse_json_object,
    read_excel_rows,
    write_excel_rows,
)


JsonDict = dict[str, Any]
ProgressCallback = Callable[[int, int, str], None]
LogCallback = Callable[[str], None]
CancelCallback = Callable[[], bool]


@dataclass(frozen=True)
class BatchOptions:
    """Runtime options for one batch job."""

    task_description: str = "Analyze the Excel row and fill the requested output fields."
    row_start: int | None = None
    row_end: int | None = None
    only_empty_outputs: bool = False
    autosave_every: int = 3
    request_delay_seconds: float = 0
    max_api_retries: int = 3
    max_parse_retries: int = 2
    retry_sleep_seconds: float = 2
    pdf_path_column: str = ""
    pdf_text_limit: int = 100000
    retained_columns: Sequence[str] = field(default_factory=tuple)
    save_raw_response: bool = True
    diagnostic_columns: bool = True


@dataclass
class BatchResult:
    """Summary returned after a batch run."""

    output_path: Path
    progress_path: Path | None
    total: int
    success: int = 0
    skipped: int = 0
    failed: int = 0
    cancelled: bool = False
    rows: list[JsonDict] = field(default_factory=list)


class BatchRowError(Exception):
    """Row-level failure that can carry the last raw model response."""

    def __init__(self, message: str, raw_response: str = "") -> None:
        super().__init__(message)
        self.raw_response = raw_response


def read_excel_columns(path: str | Path, sheet_name: str | None = None) -> list[str]:
    """Return the header row from an Excel workbook or CSV file."""

    source_path = Path(path)
    rows = read_excel_rows(path, sheet_name=sheet_name)
    if rows:
        return list(rows[0].keys())
    if source_path.suffix.lower() == ".csv":
        import csv

        encodings = ("utf-8-sig", "gb18030")
        for encoding in encodings:
            try:
                with source_path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.reader(handle)
                    return [str(value).strip() for value in next(reader) if str(value).strip()]
            except (UnicodeDecodeError, StopIteration):
                continue
        return []

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Excel headers.") from exc

    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    return [str(value).strip() for value in headers if value is not None and str(value).strip()]


def default_output_path(input_path: str | Path) -> Path:
    """Create a timestamped output path beside the input workbook."""

    source = Path(input_path)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return source.with_name(f"{source.stem}_paper2excel_{stamp}.xlsx")


def run_batch(
    input_path: str | Path,
    output_path: str | Path,
    column_mappings: Sequence[ColumnMapping | Mapping[str, Any]],
    output_fields: Sequence[OutputField | Mapping[str, Any]],
    settings: ChatCompletionSettings | None = None,
    chat: ChatCallable | None = None,
    sheet_name: str | None = None,
    options: BatchOptions | None = None,
    on_progress: ProgressCallback | None = None,
    on_log: LogCallback | None = None,
    should_cancel: CancelCallback | None = None,
) -> BatchResult:
    """Run one Excel-to-Excel AI batch job.

    The function is GUI-safe: it reports progress through callbacks, never exits
    the process, and records row failures into diagnostic columns instead of
    stopping the entire batch.
    """

    opts = options or BatchOptions()
    if chat is None:
        if settings is None:
            raise ValueError("Either settings or chat must be provided.")
        chat = lambda messages: call_chat_completion(messages, settings)

    fields = [_as_output_field(field) for field in output_fields]
    mappings = [_as_column_mapping(mapping) for mapping in column_mappings]
    required_keys = [field.name for field in fields if field.required]
    source_rows = read_excel_rows(input_path, sheet_name=sheet_name)
    indexed_rows = _select_rows(source_rows, fields, opts)
    total = len(indexed_rows)
    output_rows = [dict(row) for row in source_rows]

    out_path = Path(output_path)
    progress_path = out_path.with_name(f"{out_path.stem}_progress.xlsx")
    result = BatchResult(output_path=out_path, progress_path=progress_path, total=total, rows=output_rows)

    _log(on_log, f"Loaded {len(source_rows)} rows; selected {total} rows.")
    output_columns = _build_output_columns(source_rows, fields, opts)

    for position, row_index in enumerate(indexed_rows, start=1):
        if should_cancel and should_cancel():
            result.cancelled = True
            _log(on_log, "Cancelled by user; saving progress.")
            break

        row = output_rows[row_index]
        excel_row_number = row_index + 2
        title = _first_non_empty(row, ["标题", "Title", "title", "论文标题"]) or f"Excel row {excel_row_number}"
        _log(on_log, f"[{position}/{total}] Processing row {excel_row_number}: {str(title)[:80]}")
        if on_progress:
            on_progress(position - 1, total, f"Processing row {excel_row_number}")

        try:
            working_row = _row_with_pdf_content(row, opts, on_log)
            parsed, raw_response = _call_parse_capture(
                chat=chat,
                messages=build_json_prompt(working_row, mappings, fields, opts.task_description),
                required_keys=required_keys,
                max_api_retries=opts.max_api_retries,
                max_parse_retries=opts.max_parse_retries,
                retry_sleep_seconds=opts.retry_sleep_seconds,
            )
            for field_item in fields:
                row[field_item.name] = _normalize_field_value(parsed.get(field_item.name), field_item)
            _write_diagnostics(row, opts, "success", "", settings, raw_response)
            result.success += 1
        except Exception as exc:
            _log(on_log, f"  Failed: {exc}")
            for field_item in fields:
                row.setdefault(field_item.name, "")
            _write_diagnostics(row, opts, "failed", str(exc), settings, getattr(exc, "raw_response", ""))
            result.failed += 1

        if on_progress:
            on_progress(position, total, f"Finished row {excel_row_number}")

        if opts.autosave_every > 0 and position % opts.autosave_every == 0:
            _safe_write(progress_path, output_rows, output_columns, on_log)

        if opts.request_delay_seconds > 0 and position < total:
            time.sleep(opts.request_delay_seconds)

    unprocessed = total - result.success - result.failed
    if result.cancelled:
        result.skipped += max(0, unprocessed)
    _safe_write(out_path, output_rows, output_columns, on_log)
    result.rows = output_rows
    return result


def _call_parse_capture(
    chat: ChatCallable,
    messages: list[JsonDict],
    required_keys: Sequence[str],
    max_api_retries: int,
    max_parse_retries: int,
    retry_sleep_seconds: float,
) -> tuple[JsonDict, str]:
    current_messages = list(messages)
    last_response = ""
    last_error: Exception | None = None

    for parse_attempt in range(max_parse_retries + 1):
        for api_attempt in range(max_api_retries):
            try:
                last_response = chat(current_messages)
                break
            except ChatCompletionError as exc:
                last_error = exc
                if api_attempt >= max_api_retries - 1:
                    raise
                time.sleep(retry_sleep_seconds * (api_attempt + 1))
        try:
            return parse_json_object(last_response, required_keys=required_keys), last_response
        except JsonParseError as exc:
            last_error = exc
            if parse_attempt >= max_parse_retries:
                break
            current_messages = [
                *messages,
                {"role": "assistant", "content": last_response},
                {
                    "role": "user",
                    "content": (
                        "The previous response could not be parsed as the required JSON object: "
                        f"{exc}. Return only one corrected JSON object with the required keys."
                    ),
                },
            ]
            time.sleep(retry_sleep_seconds)

    raise BatchRowError(f"Could not parse valid JSON after retries: {last_error}", last_response)


def _row_with_pdf_content(row: JsonDict, options: BatchOptions, on_log: LogCallback | None) -> JsonDict:
    if not options.pdf_path_column:
        return row
    value = row.get(options.pdf_path_column)
    if value is None or str(value).strip() == "":
        return row

    for candidate in str(value).split(";"):
        path = candidate.strip().strip('"')
        if not path.lower().endswith(".pdf"):
            continue
        pdf_path = Path(path)
        if not pdf_path.exists():
            _log(on_log, f"  PDF not found: {pdf_path}")
            continue
        text = extract_pdf_text(pdf_path, limit=options.pdf_text_limit)
        if text:
            enriched = dict(row)
            enriched["PDF Content"] = text
            _log(on_log, f"  PDF text extracted: {len(text)} chars")
            return enriched
    return row


def extract_pdf_text(path: str | Path, limit: int = 10000) -> str:
    """Extract text from a PDF with PyMuPDF when available."""

    try:
        import fitz
    except ImportError:
        return ""

    parts: list[str] = []
    char_count = 0
    try:
        with fitz.open(path) as document:
            for page in document:
                text = page.get_text("text")
                if not text:
                    continue
                remaining = max(0, limit - char_count)
                if remaining <= 0:
                    break
                parts.append(text[:remaining])
                char_count += len(parts[-1])
                if char_count >= limit:
                    break
    except Exception:
        return ""
    return "\n".join(parts).strip()


def _select_rows(rows: Sequence[JsonDict], fields: Sequence[OutputField], options: BatchOptions) -> list[int]:
    start = max(1, options.row_start or 1)
    end = min(len(rows), options.row_end or len(rows))
    if end < start:
        return []

    selected: list[int] = []
    for row_number in range(start, end + 1):
        index = row_number - 1
        if options.only_empty_outputs and any(str(rows[index].get(field.name, "")).strip() for field in fields):
            continue
        selected.append(index)
    return selected


def _build_output_columns(rows: Sequence[JsonDict], fields: Sequence[OutputField], options: BatchOptions) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    if options.retained_columns:
        for key in options.retained_columns:
            if key not in seen:
                columns.append(key)
                seen.add(key)
    else:
        for row in rows:
            for key in row:
                if key not in seen:
                    columns.append(key)
                    seen.add(key)
    for field_item in fields:
        if field_item.name not in seen:
            columns.append(field_item.name)
            seen.add(field_item.name)
    if options.diagnostic_columns:
        for key in ["_status", "_error", "_model", "_processed_at"]:
            if key not in seen:
                columns.append(key)
                seen.add(key)
    if options.save_raw_response and "_raw_response" not in seen:
        columns.append("_raw_response")
    return columns


def _write_diagnostics(
    row: JsonDict,
    options: BatchOptions,
    status: str,
    error: str,
    settings: ChatCompletionSettings | None,
    raw_response: str,
) -> None:
    if options.diagnostic_columns:
        row["_status"] = status
        row["_error"] = error
        row["_model"] = settings.model if settings else "custom"
        row["_processed_at"] = datetime.now().isoformat(timespec="seconds")
    if options.save_raw_response:
        row["_raw_response"] = raw_response[:30000]


def _safe_write(path: Path, rows: Sequence[Mapping[str, Any]], columns: Sequence[str], on_log: LogCallback | None) -> None:
    try:
        write_excel_rows(path, rows, columns=columns)
        _log(on_log, f"Saved: {path}")
    except PermissionError:
        fallback = path.with_name(f"{path.stem}_{datetime.now().strftime('%H%M%S')}{path.suffix}")
        write_excel_rows(fallback, rows, columns=columns)
        _log(on_log, f"Target file was locked. Saved fallback: {fallback}")


def _normalize_field_value(value: Any, field: OutputField) -> Any:
    if value is None:
        return None
    field_type = field.field_type.lower()
    if field_type == "number":
        if isinstance(value, (int, float)):
            return value
        try:
            return float(str(value).strip())
        except ValueError:
            return value
    if field_type == "boolean":
        if isinstance(value, bool):
            return value
        lowered = str(value).strip().lower()
        if lowered in {"true", "yes", "y", "1", "是"}:
            return True
        if lowered in {"false", "no", "n", "0", "否"}:
            return False
        return value
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _as_column_mapping(value: ColumnMapping | Mapping[str, Any]) -> ColumnMapping:
    if isinstance(value, ColumnMapping):
        return value
    return ColumnMapping(
        source=str(value["source"]),
        alias=value.get("alias"),
        description=str(value.get("description", "")),
    )


def _as_output_field(value: OutputField | Mapping[str, Any]) -> OutputField:
    if isinstance(value, OutputField):
        return value
    return OutputField(
        name=str(value["name"]),
        description=str(value.get("description", "")),
        field_type=str(value.get("field_type", value.get("type", "string"))),
        required=bool(value.get("required", True)),
        examples=tuple(value.get("examples", ())),
    )


def _first_non_empty(row: Mapping[str, Any], keys: Sequence[str]) -> Any:
    for key in keys:
        value = row.get(key)
        if value is not None and str(value).strip():
            return value
    return None


def _log(callback: LogCallback | None, message: str) -> None:
    if callback:
        callback(message)
