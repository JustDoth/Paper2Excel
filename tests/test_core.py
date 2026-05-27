from pathlib import Path

from openpyxl import Workbook, load_workbook

from paper2excel.core import (
    ChatCompletionSettings,
    ColumnMapping,
    OutputField,
    _post_json_with_urllib,
    build_chat_completions_url,
    build_json_prompt,
    normalize_proxy_url,
    parse_json_object,
    process_excel,
    read_excel_rows,
    write_excel_rows,
)


def test_read_excel_rows(tmp_path: Path) -> None:
    path = tmp_path / "input.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Title", "Abstract"])
    sheet.append(["A paper", "About cities"])
    workbook.save(path)

    assert read_excel_rows(path) == [{"Title": "A paper", "Abstract": "About cities"}]


def test_read_csv_rows(tmp_path: Path) -> None:
    path = tmp_path / "input.csv"
    path.write_text("Title,Abstract\nA paper,About cities\n", encoding="utf-8-sig")

    assert read_excel_rows(path) == [{"Title": "A paper", "Abstract": "About cities"}]


def test_build_json_prompt_contains_structured_contract() -> None:
    messages = build_json_prompt(
        {"Title": "A paper"},
        [ColumnMapping("Title", alias="paper_title", description="论文标题")],
        [OutputField("topic", "研究主题")],
    )

    assert messages[0]["role"] == "system"
    assert "paper_title" in messages[1]["content"]
    assert "topic" in messages[1]["content"]


def test_parse_json_object_accepts_fenced_json() -> None:
    parsed = parse_json_object('```json\n{"topic": "urban growth"}\n```', required_keys=["topic"])

    assert parsed == {"topic": "urban growth"}


def test_build_chat_completions_url_accepts_base_or_full_endpoint() -> None:
    assert build_chat_completions_url("https://api.openai.com/v1") == "https://api.openai.com/v1/chat/completions"
    assert (
        build_chat_completions_url("https://api.openai.com/v1/chat/completions")
        == "https://api.openai.com/v1/chat/completions"
    )


def test_normalize_proxy_url_accepts_host_port() -> None:
    assert normalize_proxy_url("127.0.0.1:7897") == "http://127.0.0.1:7897"
    assert normalize_proxy_url("http://127.0.0.1:7897") == "http://127.0.0.1:7897"


def test_urllib_fallback_reports_original_requests_import_error() -> None:
    try:
        _post_json_with_urllib(
            "http://127.0.0.1:1/chat/completions",
            {"Content-Type": "application/json"},
            {"model": "x", "messages": []},
            ChatCompletionSettings(api_key="", model="x", timeout_seconds=1),
            ModuleNotFoundError("No module named 'requests'"),
        )
    except Exception as exc:
        assert "Requests import failed first" in str(exc)


def test_process_excel_retries_invalid_json(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Title", "Abstract"])
    sheet.append(["A paper", "About cities"])
    workbook.save(input_path)

    responses = iter(["not json", '{"topic": "urban growth", "method": "GIS"}'])

    rows = process_excel(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[
            {"source": "Title", "alias": "title", "description": "论文标题"},
            {"source": "Abstract", "alias": "abstract", "description": "摘要"},
        ],
        output_fields=[
            {"name": "topic", "description": "研究主题"},
            {"name": "method", "description": "方法"},
        ],
        chat=lambda _messages: next(responses),
        retry_sleep_seconds=0,
    )

    assert rows[0]["topic"] == "urban growth"
    saved = load_workbook(output_path)
    assert saved.active["C1"].value == "topic"
    assert saved.active["C2"].value == "urban growth"


def test_file_attachments_column_gets_hyperlink(tmp_path: Path) -> None:
    attachment = tmp_path / "paper.pdf"
    attachment.write_text("fake pdf", encoding="utf-8")
    output = tmp_path / "output.xlsx"

    write_excel_rows(
        output,
        [{"Title": "A paper", "File Attachments": str(attachment)}],
        columns=["Title", "File Attachments"],
    )

    workbook = load_workbook(output)
    cell = workbook.active["B2"]
    assert cell.hyperlink is not None
    assert "paper.pdf" in cell.hyperlink.target
