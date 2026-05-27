from pathlib import Path

from openpyxl import Workbook, load_workbook

from paper2excel.runner import BatchOptions, read_excel_columns, run_batch


def _make_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["标题", "摘要"])
    sheet.append(["城市韧性论文", "讨论人口变化和灾害恢复。"])
    sheet.append(["空摘要论文", ""])
    workbook.save(path)


def test_run_batch_writes_diagnostics_and_model_fields(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_workbook(input_path)

    rows = run_batch(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[{"source": "标题"}, {"source": "摘要"}],
        output_fields=[{"name": "结论", "description": "总结结论"}],
        chat=lambda _messages: '{"结论": "支持人口韧性分析"}',
        options=BatchOptions(row_start=1, row_end=1, retry_sleep_seconds=0),
    ).rows

    assert rows[0]["结论"] == "支持人口韧性分析"
    assert rows[0]["_status"] == "success"
    workbook = load_workbook(output_path)
    headers = [cell.value for cell in workbook.active[1]]
    assert "结论" in headers
    assert "_status" in headers


def test_run_batch_records_failed_row_without_stopping(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_workbook(input_path)
    calls = iter(['{"结论": "ok"}', "not json", "still not json", "bad again"])

    result = run_batch(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[{"source": "标题"}],
        output_fields=[{"name": "结论", "description": "总结结论"}],
        chat=lambda _messages: next(calls),
        options=BatchOptions(row_start=1, row_end=2, max_parse_retries=2, retry_sleep_seconds=0),
    )

    assert result.success == 1
    assert result.failed == 1
    assert result.rows[1]["_status"] == "failed"
    assert result.rows[1]["_raw_response"] == "bad again"


def test_run_batch_converts_number_fields(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_workbook(input_path)

    result = run_batch(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[{"source": "标题"}],
        output_fields=[{"name": "Confidence", "description": "置信度", "field_type": "number"}],
        chat=lambda _messages: '{"Confidence": "0.83"}',
        options=BatchOptions(row_start=1, row_end=1, retry_sleep_seconds=0),
    )

    assert result.rows[0]["Confidence"] == 0.83


def test_read_columns_from_csv_without_data_rows(tmp_path: Path) -> None:
    path = tmp_path / "empty.csv"
    path.write_text("标题,摘要\n", encoding="utf-8-sig")

    assert read_excel_columns(path) == ["标题", "摘要"]


def test_run_batch_keeps_only_retained_source_columns(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_workbook(input_path)

    run_batch(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[{"source": "标题"}, {"source": "摘要"}],
        output_fields=[{"name": "结论", "description": "总结结论"}],
        chat=lambda _messages: '{"结论": "ok"}',
        options=BatchOptions(row_start=1, row_end=1, retained_columns=("标题",), retry_sleep_seconds=0),
    )

    workbook = load_workbook(output_path)
    headers = [cell.value for cell in workbook.active[1]]
    assert "标题" in headers
    assert "摘要" not in headers
    assert "结论" in headers


def test_run_batch_preserves_retained_and_output_field_order(tmp_path: Path) -> None:
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _make_workbook(input_path)

    run_batch(
        input_path=input_path,
        output_path=output_path,
        column_mappings=[{"source": "标题"}, {"source": "摘要"}],
        output_fields=[
            {"name": "Second", "description": "second"},
            {"name": "First", "description": "first"},
        ],
        chat=lambda _messages: '{"Second": "2", "First": "1"}',
        options=BatchOptions(row_start=1, row_end=1, retained_columns=("摘要", "标题"), retry_sleep_seconds=0),
    )

    workbook = load_workbook(output_path)
    headers = [cell.value for cell in workbook.active[1]]
    assert headers[:7] == ["摘要", "标题", "Second", "First", "_status", "_error", "_model"]
